import chromedriver_autoinstaller
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
import os
import openpyxl
import datetime
from openpyxl.drawing.image import Image
import urllib.request as req
import re
import PIL

if not os.path.exists("./네이버_쇼핑"):
    os.mkdir("./네이버_쇼핑")

if not os.path.exists("./네이버_쇼핑.xlsx"):
    openpyxl.Workbook().save("./네이버_쇼핑.xlsx")

book = openpyxl.load_workbook("./네이버_쇼핑.xlsx")

if "Sheet" in book.sheetnames:
    book.remove(book["Sheet"])

sheet = book.create_sheet()

sheet.column_dimensions["A"].width = 14.50
sheet.column_dimensions["B"].width = 90
sheet.column_dimensions["C"].width = 10
sheet.column_dimensions["D"].width = 10
sheet.column_dimensions["E"].width = 10
sheet.column_dimensions["F"].width = 10

item = input("원하는 상품을 입력하세요 : >> ")

chrome_path = chromedriver_autoinstaller.install()
browser = webdriver.Chrome(chrome_path)


if not os.path.exists("./네이버_쇼핑/{}".format(item)):
    os.mkdir("./네이버_쇼핑/{}".format(item))

if not os.path.exists("./네이버_쇼핑/{}/{}".format(item, "resize")):
    os.mkdir("./네이버_쇼핑/{}/{}".format(item, "resize"))

now = datetime.datetime.now()
sheet.title = "{}년 {}월 {}일 {}시 {}분 {}초 {}".format(now.year,now.month,now.day,now.hour,now.minute,now.second,item)

page_num = 1
row_num = 1

while True:
    browser.get("https://search.shopping.naver.com/search/all?frm=NVSHATC&origQuery={}&pagingIndex={}&pagingSize=20&productSet=total&query={}&sort=rel&timestamp=&viewType=list".format(item, page_num, item))
    time.sleep(1)
    for _ in range(3):
        browser.find_element_by_css_selector("html").send_keys(Keys.END)
        time.sleep(0.5)

    # 상품
    product = browser.find_elements_by_css_selector("li.basicList_item__2XT81")

    for i in product:
        name = i.find_element_by_css_selector("div.basicList_title__3P9Q7 > a")
        price = i.find_element_by_css_selector("span.price_num__2WUXn")


        try:
            review = i.find_element_by_css_selector("div.basicList_etc_box__1Jzg6 > a:nth-child(1)> em.basicList_num__1yXM9").text
        except:
            review = "0"
        try:
            buy_num = i.find_element_by_css_selector("div.basicList_etc_box__1Jzg6 > a:nth-child(2)> em.basicList_num__1yXM9").text
        except:
            buy_num = "0"

        try:
            img = i.find_element_by_css_selector("ul.list_basis div.thumbnail_thumb_wrap__1pEkS._wrapper img")

        except:
            print("이미지가 없습니다.")
            print()
            pass


        img_url = img.get_attribute("src")
        img_dir = "./네이버_쇼핑/{}/{}.png".format(item,re.sub("[\\\/:*?\"<>\|]", " ", name.text))
        req.urlretrieve(img_url,img_dir)
        im = PIL.Image.open(img_dir).convert('RGB')
        im_small = im.resize((120,120),PIL.Image.ANTIALIAS)
        im_small_dir = "./네이버_쇼핑/{}/{}/{}.png".format(item,"resize",re.sub("[\\\/:*?\"<>\|]", " ", name.text))
        im_small.save(im_small_dir,format("png"),quality = 100)
        img_for_excel = Image(im_small_dir)
        sheet.add_image(img_for_excel, "A{}".format(row_num))
        sheet.cell(row=row_num, column=2).value = name.text
        sheet.cell(row=row_num, column=3).value = price.text
        sheet.cell(row=row_num, column=4).value = review
        sheet.cell(row=row_num, column=5).value = buy_num
        sheet.cell(row=row_num, column=6).value = name.get_attribute("href")
        sheet.row_dimensions[row_num].height = 80
        print("제품명 : {} ".format(name.text))
        print("가격 : {}".format(price.text))
        print("리뷰 : {}건".format(review))
        print("구매건수 : {}건".format(buy_num))
        print("링크 : {} ".format(name.get_attribute("href")))
        print()
        book.save("./네이버_쇼핑.xlsx")
        row_num += 1
    page_num += 1

    if page_num == 4:
        browser.close()
        break
        # browser = webdriver.Chrome("./chromedriver")
        # browser.implicitly_wait(3)